"""
Scrape product listings from koreanshopbd.com search results using Playwright.
Extracts product name, price, original price, image URL, and product URL.
"""

import re
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

BASE_URL = "https://koreanshopbd.com"

HEADERS = ["name", "price", "origin_price", "image_url", "product_url"]


def write_xlsx(products: list[dict[str, Any]], output_path: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    for col, h in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for row_idx, p in enumerate(products, 2):
        ws.cell(row=row_idx, column=1, value=p.get("name", ""))
        ws.cell(row=row_idx, column=2, value=p.get("price", ""))
        ws.cell(row=row_idx, column=3, value=p.get("origin_price", ""))
        ws.cell(row=row_idx, column=4, value=p.get("image_url", ""))
        ws.cell(row=row_idx, column=5, value=p.get("product_url", ""))

    for col in (2, 3):
        ws.column_dimensions[chr(64 + col)].width = 14
    ws.column_dimensions["A"].width = 50
    ws.column_dimensions["D"].width = 60
    ws.column_dimensions["E"].width = 60

    wb.save(output_path)
    print(f"  Saved {len(products)} products to: {output_path}")


async def scrape_search_page(page, search_query: str, output_file: str | None = None) -> list[dict[str, Any]]:
    """
    Scrape product listings from a Korean Shop BD search results page.

    Args:
        page: A Playwright page instance (already navigated or about to navigate).
        search_query: The search term to look up.
        output_file: Optional JSON file path to save results.

    Returns:
        List of product dicts with keys: name, price, origin_price, image_url, product_url.
    """
    search_url = f"{BASE_URL}/search?searchQuery={search_query}"
    print(f"  Navigating to: {search_url}")
    await page.goto(search_url, wait_until="networkidle", timeout=30000)
    await page.wait_for_timeout(2000)

    products = await page.evaluate("""
        () => {
            const results = [];

            // Select all product card anchor tags inside the grid
            const cards = document.querySelectorAll('a[href^="/product/"]');
            if (cards.length === 0) {
                // Fallback: try selecting any matching grid layout
                const grid = document.querySelector('.grid.gap-3');
                if (grid) {
                    return [{ error: 'grid found but no product cards matched a[href^="/product/"]' }];
                }
                return [{ error: 'no product grid found' }];
            }

            for (const card of cards) {
                // Product name from the <p> tag inside .px-3
                const nameEl = card.querySelector('.px-3 > p');
                const name = nameEl ? nameEl.textContent.trim() : '';

                // Image URL from the <img> inside the card
                const imgEl = card.querySelector('img[src]');
                const image_url = imgEl ? imgEl.getAttribute('src') : '';

                // Price div - contains current price and optional original (strikethrough) price
                const priceDiv = card.querySelector('.text-lg.font-semibold.text-secondary');
                let price = '';
                let origin_price = '';

                if (priceDiv) {
                    const fullText = priceDiv.textContent.trim();

                    // Current price is the first price (before any strikethrough span)
                    const strikeSpan = priceDiv.querySelector('span.line-through, .line-through');
                    if (strikeSpan) {
                        origin_price = strikeSpan.textContent.trim();
                        // Current price is whatever remains when we remove the strikethrough text
                        const strikeText = strikeSpan.textContent.trim();
                        const mainText = fullText.replace(strikeText, '').trim();
                        price = mainText;
                    } else {
                        price = fullText;
                    }
                }

                // Clean price strings: remove currency symbols, keep only digits and dots
                const cleanPrice = (p) => {
                    return p.replace(/[^\\d.]/g, '').trim();
                };

                // Product URL
                const href = card.getAttribute('href');
                const product_url = href ? (href.startsWith('http') ? href : window.location.origin + href) : '';

                results.push({
                    name: name,
                    price: cleanPrice(price),
                    origin_price: origin_price ? cleanPrice(origin_price) : '',
                    image_url: image_url,
                    product_url: product_url,
                });
            }

            return results;
        }
    """)

    if not products:
        print("  No products found.")
        return []

    # Filter out error entries
    valid = [p for p in products if "error" not in p]

    print(f"  Found {len(valid)} product(s):")
    for p in valid:
        print(f"    - {p['name']} | ৳{p['price']}" + (f" (was ৳{p['origin_price']})" if p['origin_price'] else ""))

    if output_file and valid:
        out_path = Path(output_file)
        out_path.parent.mkdir(parents=True, exist_ok=True)
        if out_path.suffix.lower() == ".xlsx":
            write_xlsx(valid, str(out_path))
        else:
            import json
            out_path.write_text(json.dumps(valid, indent=2, ensure_ascii=False), encoding="utf-8")
            print(f"  Saved to: {out_path}")

    return valid


async def main():
    import argparse

    parser = argparse.ArgumentParser(description="Scrape koreanshopbd.com search results")
    parser.add_argument("--query", default="aplb", help="Search query (default: aplb)")
    parser.add_argument("--output", default=None, help="Output file path (.xlsx or .json, default: {query}_products.xlsx)")
    parser.add_argument("--headless", action="store_true", help="Run headless")
    args = parser.parse_args()

    if not args.output:
        args.output = f"{args.query}_products.xlsx"

    from playwright.async_api import async_playwright

    async with async_playwright() as pw:
        browser = await pw.chromium.launch(headless=args.headless, channel="chrome")
        page = await browser.new_page(viewport={"width": 1920, "height": 1080})
        try:
            await scrape_search_page(page, args.query, args.output)
        finally:
            await browser.close()


if __name__ == "__main__":
    import asyncio
    asyncio.run(main())
